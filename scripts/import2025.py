"""
import2025.py — Parse the 2025 Excel spreadsheet and produce:
  data/2025-import.json   — full season data in the app's Firebase format
  src/constants/league_2025.js — season constants for the app

Usage:
  python3 scripts/import2025.py

Requirements:
  pip install xlrd          # for .xls files
  pip install openpyxl      # if you convert the file to .xlsx first
"""

import json
import os
import sys

XLS = os.path.expanduser(
    "~/Desktop/Golf League Master v2.2025 - Copy.xls"
)

print(f"Loading {XLS} …")

# Try openpyxl first (works if file is actually xlsx/xlsm despite extension),
# then fall back to xlrd for true .xls binary format.
try:
    import openpyxl
    wb_raw = openpyxl.load_workbook(XLS, read_only=True, keep_vba=False, data_only=True)

    def get_sheet(name):
        return wb_raw[name]

    def iter_rows(ws, min_row, max_row):
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
            yield list(row)

    def sheet_names():
        return wb_raw.sheetnames

    print("  (opened with openpyxl)")

except Exception as e_openpyxl:
    try:
        import xlrd
        wb_raw = xlrd.open_workbook(XLS)

        class _XlrdSheet:
            def __init__(self, ws):
                self._ws = ws

        def get_sheet(name):
            return wb_raw.sheet_by_name(name)

        def iter_rows(ws, min_row, max_row):
            for r in range(min_row - 1, min(max_row, ws.nrows)):
                row = ws.row_values(r)
                # xlrd returns floats for integers — convert where needed
                out = []
                for v in row:
                    if isinstance(v, float) and v == int(v):
                        out.append(int(v))
                    elif v == "":
                        out.append(None)
                    else:
                        out.append(v)
                yield out

        def sheet_names():
            return wb_raw.sheet_names()

        print("  (opened with xlrd)")

    except Exception as e_xlrd:
        print(f"ERROR: Could not open {XLS}")
        print(f"  openpyxl: {e_openpyxl}")
        print(f"  xlrd:     {e_xlrd}")
        print("\nFix: pip install xlrd  OR  save the file as .xlsx and update the XLS path above.")
        sys.exit(1)


# ── 1. Teams + player IDs from Main sheet ────────────────────────────────────
ws_main = get_sheet("Main")
teams       = {}          # tid  → {p1, p2, name}
pid_to_team = {}          # pid  → (tid, pi)

for r in iter_rows(ws_main, min_row=10, max_row=30):
    if len(r) < 9:
        continue
    tid, p1id, p1, p2id, p2 = r[4], r[5], r[6], r[7], r[8]
    if not tid:
        continue
    tid = int(tid)
    p1  = str(p1).strip() if p1 else ""
    p2  = str(p2).strip() if p2 else ""
    teams[tid] = {
        "p1":   p1,
        "p2":   p2,
        "name": f"{p1.split()[-1]} - {p2.split()[-1]}",
    }
    if p1id:
        pid_to_team[int(p1id)] = (tid, 0)
    if p2id:
        pid_to_team[int(p2id)] = (tid, 1)

# pid lookup by (tid, pi)
team_pids = {}   # (tid, pi) → pid
for pid, (tid, pi) in pid_to_team.items():
    team_pids[(tid, pi)] = pid

print(f"  {len(teams)} teams, {len(pid_to_team)} players")


# ── 2. Handicaps per player per week from Summary sheet ──────────────────────
#    col 44 (0-indexed) = HCP going into Week 1, col 45 = Week 2, … col 61 = Week 18
ws_sum = get_sheet("Summary")
player_hcp = {}   # pid → list of 19 values (index 0 unused; index 1-18 = weeks)

for r in iter_rows(ws_sum, min_row=4, max_row=40):
    if not r or not r[0]:
        continue
    pid = int(r[0])
    player_hcp[pid] = [None] + [
        (int(r[44 + w]) if r[44 + w] is not None else None)
        if len(r) > 44 + w else None
        for w in range(18)
    ]


def team_hcp_for_week(tid, week):
    """Return [pi0_hcp, pi1_hcp] for a team in a given week (1-indexed)."""
    hcps = [0, 0]
    for pid, (t, pi) in pid_to_team.items():
        if t == tid:
            raw = player_hcp.get(pid, [None] * 19)[week]
            hcps[pi] = int(raw) if raw is not None else 0
    return hcps


# ── 3. Parse each week sheet ─────────────────────────────────────────────────
results  = {}   # week → { matchKey: record }
schedule = {}   # week → [[tlow, thigh], …]

def detect_pi(raw_row, tid):
    """Read player ID from a raw row and return the correct pi (0 or 1)."""
    for col in [1, 2, 3, 6, 28, 30]:
        if col >= len(raw_row):
            continue
        val = raw_row[col]
        if val and isinstance(val, (int, float)):
            pid = int(val)
            if pid in pid_to_team:
                t, pi = pid_to_team[pid]
                if t == tid:
                    return pi
    return None


def raw_scores(row):
    """Cols 7-15 (0-indexed) → 9 hole scores as integers."""
    out = []
    for v in row[7:16]:
        if v is None or v == "" or v == "-":
            out.append(0)
        else:
            try:
                out.append(int(float(v)))
            except (TypeError, ValueError):
                out.append(0)
    return out


def player_type(row):
    """Determine type from col4 (Sub), col5 (Phantom), or all-zero scores."""
    if row[4] == "Y":
        return "sub"
    if row[5] == "Y":
        return "phantom"
    scores = row[7:16]
    if all(v is None or v == 0 or v == "" or v == "-" for v in scores):
        return "phantom"
    return "normal"


all_sheets = sheet_names()

for wk in range(1, 22):   # 2025 goes up to week 21
    sheet_name = f"Week{wk}"
    if sheet_name not in all_sheets:
        continue

    ws   = get_sheet(sheet_name)
    rows = list(iter_rows(ws, min_row=1, max_row=9999))
    results[wk]  = {}
    schedule[wk] = []

    i = 0
    while i < len(rows):
        row = rows[i]
        val = row[0] if row else None

        # Match blocks begin where col A is an integer match number 1-9
        if not (isinstance(val, int) and 1 <= val <= 9):
            i += 1
            continue

        # Each match = 12 rows (4 players × 3 sub-rows: Raw, Handicap, Adjusted)
        block = rows[i: i + 12]
        if len(block) < 12:
            break

        # Team IDs: col 29 of the Adjusted row of the 2nd player of each pair
        tid1_raw = block[5][29] if len(block[5]) > 29 else None
        tid2_raw = block[11][29] if len(block[11]) > 29 else None
        if not tid1_raw or not tid2_raw:
            i += 12
            continue
        tid1, tid2 = int(tid1_raw), int(tid2_raw)

        pi_a = detect_pi(block[0], tid1)
        if pi_a is None: pi_a = 0
        pi_b = 1 - pi_a

        pi_c = detect_pi(block[6], tid2)
        if pi_c is None: pi_c = 0
        pi_d = 1 - pi_c

        s = [raw_scores(block[p * 3]) for p in range(4)]
        t = [player_type(block[p * 3]) for p in range(4)]

        t1s = [None, None]; t1t = [None, None]
        t2s = [None, None]; t2t = [None, None]
        t1s[pi_a] = s[0];   t1t[pi_a] = t[0]
        t1s[pi_b] = s[1];   t1t[pi_b] = t[1]
        t2s[pi_c] = s[2];   t2t[pi_c] = t[2]
        t2s[pi_d] = s[3];   t2t[pi_d] = t[3]

        t1s = [x or [0] * 9 for x in t1s]
        t2s = [x or [0] * 9 for x in t2s]
        t1t = [x or "phantom" for x in t1t]
        t2t = [x or "phantom" for x in t2t]

        tlow  = min(tid1, tid2)
        thigh = max(tid1, tid2)
        mk    = f"{wk}-{tlow}-{thigh}"

        hcp1 = team_hcp_for_week(tid1, wk)
        hcp2 = team_hcp_for_week(tid2, wk)

        if tid1 == tlow:
            rec = dict(t1scores=t1s, t2scores=t2s, t1types=t1t, t2types=t2t,
                       hcpSnapshot={tlow: hcp1, thigh: hcp2})
        else:
            rec = dict(t1scores=t2s, t2scores=t1s, t1types=t2t, t2types=t1t,
                       hcpSnapshot={tlow: hcp2, thigh: hcp1})

        results[wk][mk] = rec

        pair = sorted([tid1, tid2])
        if pair not in schedule[wk]:
            schedule[wk].append(pair)

        i += 12


# ── 4. Build hcpOverrides from Summary sheet handicaps ───────────────────────
hcp_overrides = {}
for pid, hcps in player_hcp.items():
    if pid not in pid_to_team:
        continue
    tid, pi = pid_to_team[pid]
    for week in range(1, 22):
        val = hcps[week] if len(hcps) > week else None
        if val is not None:
            hcp_overrides[f"{tid}-{pi}-{week}"] = int(val)

# ── 5. Build defaultHcp (Week 1 handicaps) ───────────────────────────────────
default_hcp = {tid: team_hcp_for_week(tid, 1) for tid in sorted(teams)}

# ── 6. Write data/2025-import.json ───────────────────────────────────────────
output = {
    "teams":        {str(k): v for k, v in sorted(teams.items())},
    "defaultHcp":   {str(k): v for k, v in sorted(default_hcp.items())},
    "schedule":     {str(k): v for k, v in sorted(schedule.items())},
    "hcpOverrides": hcp_overrides,
    "results":      {str(k): v for k, v in sorted(results.items())},
}

os.makedirs("data", exist_ok=True)
with open("data/2025-import.json", "w") as f:
    json.dump(output, f, indent=2)

print("\n✓ data/2025-import.json written")
total = sum(len(v) for v in output["results"].values())
for wk in range(1, 22):
    n = len(output["results"].get(str(wk), {}))
    flag = "  (bye/playoff)" if n == 0 else ""
    print(f"  Week {wk:2d}: {n} matches{flag}")
print(f"  Total : {total} match records")

# ── 7. Write src/constants/league_2025.js ────────────────────────────────────
par_js  = "[4,3,4,5,4,3,4,5,4]"
si_js   = "[1,3,7,8,4,9,2,6,5]"

teams_js_lines = []
for tid, t in sorted(teams.items()):
    teams_js_lines.append(
        f'  {tid}: {{name:"{t["name"]}", p1:"{t["p1"]}", p2:"{t["p2"]}"}}'
    )
teams_js = "{\n" + ",\n".join(teams_js_lines) + "\n}"

default_hcp_js_lines = []
for tid, hcps in sorted(default_hcp.items()):
    default_hcp_js_lines.append(f"  {tid}: [{hcps[0]},{hcps[1]}]")
default_hcp_js = "{\n" + ",\n".join(default_hcp_js_lines) + "\n}"

schedule_raw_lines = []
for wk in range(1, 22):
    pairs = schedule.get(wk, [])
    if not pairs:
        continue
    pairs_js = ",".join(f"[{a},{b}]" for a, b in pairs)
    schedule_raw_lines.append(f'  [{wk}, null, {pairs_js}]')
schedule_raw_js = "[\n" + ",\n".join(schedule_raw_lines) + "\n]"

js_content = f"""// Auto-generated by scripts/import2025.py — do not edit by hand
// 2025 season constants

export const PAR  = {par_js};
export const SI   = {si_js};

export const RAINOUT_SUB = {{
  6: 0, 7: 3, 8: 2,
}};

export const TEAMS = {teams_js};

export const ALL_PLAYERS = Object.entries(TEAMS).flatMap(([tid, t]) => [
  {{tid: parseInt(tid), pi: 0, name: t.p1, team: t.name}},
  {{tid: parseInt(tid), pi: 1, name: t.p2, team: t.name}},
]);

export const DEFAULT_HCP = {default_hcp_js};

export const NEW_MEMBERS = {{}};

export const HCP_PCT = [0, 0.90, 0.90, 0.90, 0.90];
export const HCP_CAP = 2;
export const HCP_ROUNDS = 7;
export const NEW_MEMBER_HCP_PCT = 0.65;
export const PLAYOFF_START_WEEK = 18;

export const SCHEDULE_RAW = {schedule_raw_js};

export const BASE_TEE_TIMES = [
  "4:10 PM","4:20 PM","4:30 PM","4:40 PM","4:50 PM",
  "5:00 PM","5:10 PM","5:20 PM","5:30 PM",
];
export const TEE_TIME_OVERRIDES = {{}};

export function buildSchedule() {{
  const s = {{}};
  for (const [week, date, ...pairs] of SCHEDULE_RAW) {{
    s[week] = {{week, date, pairs: (pairs || []).filter(p => Array.isArray(p))}};
  }}
  return s;
}}

export const SCHEDULE = buildSchedule();

export function getTeeTimes(week) {{
  return TEE_TIME_OVERRIDES[week] || BASE_TEE_TIMES;
}}

export function isNewMember(tid, pi) {{
  return !!NEW_MEMBERS[`${{tid}}-${{pi}}`];
}}
"""

out_path = os.path.join(os.path.dirname(__file__), "..", "src", "constants", "league_2025.js")
with open(out_path, "w") as f:
    f.write(js_content)

print(f"\n✓ src/constants/league_2025.js written")
print("\nNext steps:")
print("  1. Run: node scripts/push2025.cjs  (to load data into Firebase)")
print("  2. Verify the 2025 season in the app's season switcher")
